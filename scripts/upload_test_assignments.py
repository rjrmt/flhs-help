#!/usr/bin/env python3
"""Convert a testing Excel/CSV export and upload to private.test_assignments.

Expected columns (header names may have trailing spaces):
  Local ID, Last, First, Grade, EXAM, DAY, DATE, AM/PM, ROOM

Usage:
  python3 scripts/upload_test_assignments.py "/path/to/RJ SEPTEMBER TEST UPLOAD.xlsx" --api

Requires SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY for --api.
Never put the service role key in the website.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import urllib.error
import urllib.request
from datetime import date, datetime
from pathlib import Path

STUDENT_ID_LEN = 10
ROOT = Path(__file__).resolve().parents[1]


def normalize_student_id(val: object) -> str:
    digits = re.sub(r"\D", "", str(val or ""))
    if not digits:
        return ""
    if len(digits) < STUDENT_ID_LEN:
        return digits.zfill(STUDENT_ID_LEN)
    return digits


def clean_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\ufeff", "").strip())


def pick(row: dict[str, object], *names: str) -> str:
    lower = {clean_header(k).lower(): v for k, v in row.items()}
    for name in names:
        key = name.lower()
        if key in lower and lower[key] is not None and str(lower[key]).strip() != "":
            return str(lower[key]).strip()
    return ""


def format_exam_date(val: object) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    text = str(val).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    # Excel serial as string
    if re.fullmatch(r"\d+(\.\d+)?", text):
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(float(text)).date().isoformat()
        except Exception:
            pass
    return None


def normalize_row(raw: dict[str, object]) -> dict[str, object] | None:
    student_id = normalize_student_id(pick(raw, "Local ID", "student_id", "Student ID", "ID"))
    if not student_id:
        return None
    exam_date = format_exam_date(pick(raw, "DATE", "Date", "EXAM DATE") or raw.get("DATE ") or raw.get("DATE"))
    # Also try original keys with trailing spaces
    if exam_date is None:
        for key, value in raw.items():
            if clean_header(key).upper() in {"DATE", "EXAM DATE"}:
                exam_date = format_exam_date(value)
                break
    return {
        "student_id": student_id,
        "last_name": pick(raw, "Last", "LAST", "Last Name"),
        "first_name": pick(raw, "First", "FIRST", "First Name"),
        "grade": re.sub(r"\D", "", pick(raw, "Grade", "GRADE")) or pick(raw, "Grade", "GRADE"),
        "exam": pick(raw, "EXAM", "Exam", "TEST", "Test"),
        "day": pick(raw, "DAY", "Day", "WEEKDAY"),
        "exam_date": exam_date,
        "session": pick(raw, "AM/PM", "SESSION", "Session").upper(),
        "room": pick(raw, "ROOM", "Room", "EXAM ROOM"),
    }


def read_xlsx(path: Path) -> list[dict[str, object]]:
    try:
        import openpyxl
    except ImportError as err:
        raise SystemExit("Install openpyxl: python3 -m pip install openpyxl") from err

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [clean_header(h) for h in next(rows_iter)]
    except StopIteration:
        wb.close()
        return []
    out: list[dict[str, object]] = []
    for values in rows_iter:
        raw = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if not any(v is not None and str(v).strip() for v in raw.values()):
            continue
        normalized = normalize_row(raw)
        if normalized:
            out.append(normalized)
    wb.close()
    return out


def read_csv(path: Path) -> list[dict[str, object]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        out: list[dict[str, object]] = []
        for raw in reader:
            cleaned = {clean_header(k): v for k, v in raw.items()}
            normalized = normalize_row(cleaned)
            if normalized:
                out.append(normalized)
        return out


def load_rows(path: Path) -> list[dict[str, object]]:
    if not path.is_file():
        raise SystemExit(f"File not found: {path}")
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    if suffix == ".csv":
        return read_csv(path)
    raise SystemExit(f"Unsupported file type: {suffix}")


def post_rpc(url: str, service_key: str, fn: str, payload: dict) -> object:
    request = urllib.request.Request(
        f"{url.rstrip('/')}/rest/v1/rpc/{fn}",
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        },
    )
    try:
        with urllib.request.urlopen(request) as response:
            body = response.read().decode("utf-8")
    except urllib.error.HTTPError as err:
        detail = err.read().decode("utf-8", errors="replace")
        raise SystemExit(f"Upload failed ({err.code}): {detail}") from err
    return json.loads(body)


def upload_via_api(rows: list[dict[str, object]], batch_size: int = 400) -> int:
    url = os.environ.get("SUPABASE_URL") or os.environ.get("FLHS_SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise SystemExit(
            "Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY to upload.\n"
            "The service role key must stay off the website."
        )
    # Replace in one call when possible; chunk only if payload is huge.
    if len(rows) <= batch_size:
        return int(post_rpc(url, key, "admin_replace_test_assignments", {"p_rows": rows}))

    # Large files: first chunk replaces, later chunks append via temporary upsert helper.
    # For safety, replace with full list in sequential batches by concatenating locally
    # and calling once with full JSON if under ~4MB; otherwise raise.
    payload = json.dumps(rows)
    if len(payload) > 4_500_000:
        raise SystemExit("File too large for a single replace; split the Excel sheet.")
    return int(post_rpc(url, key, "admin_replace_test_assignments", {"p_rows": rows}))


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path, help="Excel (.xlsx) or CSV testing export")
    parser.add_argument("--api", action="store_true", help="Upload via service-role RPC")
    parser.add_argument("--json", action="store_true", help="Print normalized JSON")
    parser.add_argument(
        "--csv-out",
        type=Path,
        help="Also write a local gitignored CSV (data/test-rooms.csv)",
    )
    args = parser.parse_args()

    rows = load_rows(args.source)
    students = len({row["student_id"] for row in rows})
    print(f"{len(rows)} assignments · {students} students", file=sys.stderr)

    if args.csv_out:
        args.csv_out.parent.mkdir(parents=True, exist_ok=True)
        with args.csv_out.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=[
                    "Local ID",
                    "Last",
                    "First",
                    "Grade",
                    "EXAM",
                    "DAY",
                    "DATE",
                    "AM/PM",
                    "ROOM",
                ],
            )
            writer.writeheader()
            for row in rows:
                writer.writerow(
                    {
                        "Local ID": row["student_id"],
                        "Last": row["last_name"],
                        "First": row["first_name"],
                        "Grade": row["grade"],
                        "EXAM": row["exam"],
                        "DAY": row["day"],
                        "DATE": row["exam_date"] or "",
                        "AM/PM": row["session"],
                        "ROOM": row["room"],
                    }
                )
        print(f"Wrote {args.csv_out}", file=sys.stderr)

    if args.json:
        json.dump(rows, sys.stdout, indent=2)
        sys.stdout.write("\n")
        return

    if args.api:
        uploaded = upload_via_api(rows)
        print(f"Uploaded {uploaded} assignments ({students} students)")
        return

    print("Re-run with --api to upload, or --json / --csv-out to inspect.")


if __name__ == "__main__":
    main()
